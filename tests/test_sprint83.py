from io import BytesIO

from openpyxl import load_workbook

from app.extensions import db
from app.models import Guest, Wedding
from app.services.guest_roundtrip import import_guest_rows


def test_roundtrip_import_updates_and_creates(app):
    with app.app_context():
        wedding = db.session.scalar(db.select(Wedding).limit(1))
        existing = Guest(
            wedding_id=wedding.id,
            first_name="ישראל",
            last_name="ישראלי",
            phone="0501111111",
            invited_count=2,
        )
        db.session.add(existing)
        db.session.commit()

        rows = [
            {
                "UUID": existing.uuid,
                "שם פרטי": "ישראל",
                "שם משפחה": "כהן",
                "טלפון": "0501111111",
                "אימייל": "",
                "הוזמנו": 4,
                "סטטוס": "מגיע",
            },
            {
                "שם פרטי": "רבקה",
                "שם משפחה": "לוי",
                "טלפון": "",
                "אימייל": "",
                "הוזמנו": "not-a-number",
            },
        ]
        result = import_guest_rows(wedding.id, rows, "update")
        assert result.updated == 1
        assert result.created == 1
        assert result.warning_count == 1
        assert db.session.get(Guest, existing.id).last_name == "כהן"
        new_guest = db.session.scalar(db.select(Guest).where(Guest.first_name == "רבקה"))
        assert new_guest is not None
        assert new_guest.phone is None
        assert new_guest.invited_count == 1


def test_guest_export_is_roundtrip_workbook(client):
    client.post(
        "/auth/login",
        data={"email": "admin@example.com", "password": "password123"},
        follow_redirects=True,
    )
    response = client.get("/guests/export.xlsx")
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.data), data_only=True)
    assert "מוזמנים" in workbook.sheetnames
    assert "_מערכת" in workbook.sheetnames
    headers = [cell.value for cell in workbook["מוזמנים"][1]]
    assert "UUID" in headers
    assert "שם פרטי" in headers
    assert "הערות" in headers
