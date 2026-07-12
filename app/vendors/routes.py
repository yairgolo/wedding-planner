from __future__ import annotations

from io import BytesIO
from urllib.parse import quote

from flask import Blueprint, abort, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import or_

from app.extensions import db
from app.models import AuditLog, Document, Task, Vendor, Wedding
from app.services.vendor_budget import sync_vendor_to_budget

from .forms import VendorForm

vendors_bp = Blueprint("vendors", __name__, url_prefix="/vendors")

CATEGORY_LABELS = {
    "venue": "אולם",
    "music": "DJ / מוזיקה",
    "photography": "צילום",
    "video": "וידאו",
    "dress": "שמלה",
    "suit": "חליפה",
    "beauty": "איפור ושיער",
    "rabbi": "רב / חופה",
    "attraction": "אטרקציות",
    "decor": "עיצוב",
    "transport": "הסעות",
    "other": "אחר",
}
STATUS_LABELS = {
    "considering": "בבדיקה",
    "negotiating": "במשא ומתן",
    "booked": "נסגר",
    "completed": "הושלם",
    "cancelled": "בוטל",
}


def current_wedding() -> Wedding:
    wedding = db.session.scalar(db.select(Wedding).order_by(Wedding.id).limit(1))
    if not wedding:
        abort(404)
    return wedding


def audit(wedding_id: int, vendor: Vendor, action: str, description: str) -> None:
    db.session.add(
        AuditLog(
            wedding_id=wedding_id,
            user_id=current_user.id,
            entity_type="vendor",
            entity_id=str(vendor.id or "new"),
            action=action,
            description=description,
        )
    )


@vendors_bp.get("")
@login_required
def index():
    wedding = current_wedding()
    q = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    status = request.args.get("status", "").strip()
    favorites = request.args.get("favorites", "").strip()

    stmt = db.select(Vendor).where(Vendor.wedding_id == wedding.id, Vendor.deleted_at.is_(None))
    if q:
        pattern = f"%{q}%"
        stmt = stmt.where(
            or_(
                Vendor.name.ilike(pattern),
                Vendor.contact_name.ilike(pattern),
                Vendor.phone.ilike(pattern),
                Vendor.email.ilike(pattern),
                Vendor.notes.ilike(pattern),
            )
        )
    if category:
        stmt = stmt.where(Vendor.category == category)
    if status:
        stmt = stmt.where(Vendor.status == status)
    if favorites == "1":
        stmt = stmt.where(Vendor.is_favorite.is_(True))

    vendors = db.session.scalars(
        stmt.order_by(Vendor.is_favorite.desc(), Vendor.category, Vendor.name)
    ).all()
    all_vendors = db.session.scalars(
        db.select(Vendor).where(Vendor.wedding_id == wedding.id, Vendor.deleted_at.is_(None))
    ).all()
    stats = {
        "total": len(all_vendors),
        "booked": sum(1 for vendor in all_vendors if vendor.status in {"booked", "completed"}),
        "unsigned": sum(
            1
            for vendor in all_vendors
            if vendor.status in {"booked", "completed"} and not vendor.contract_signed
        ),
        "balance": sum(
            float(vendor.balance) for vendor in all_vendors if vendor.status != "cancelled"
        ),
        "overdue": sum(1 for vendor in all_vendors if vendor.is_payment_overdue),
    }
    return render_template(
        "vendors/index.html",
        vendors=vendors,
        stats=stats,
        category_labels=CATEGORY_LABELS,
        status_labels=STATUS_LABELS,
        q=q,
    )


@vendors_bp.get("/<int:vendor_id>")
@login_required
def detail(vendor_id: int):
    wedding = current_wedding()
    vendor = db.get_or_404(Vendor, vendor_id)
    if vendor.wedding_id != wedding.id or vendor.is_deleted:
        abort(404)
    tasks = db.session.scalars(
        db.select(Task)
        .where(
            Task.wedding_id == wedding.id,
            Task.related_vendor_id == vendor.id,
            Task.deleted_at.is_(None),
        )
        .order_by(Task.status, Task.due_date)
    ).all()
    documents = db.session.scalars(
        db.select(Document)
        .where(
            Document.wedding_id == wedding.id,
            Document.vendor_id == vendor.id,
            Document.deleted_at.is_(None),
        )
        .order_by(Document.created_at.desc())
    ).all()
    activity = db.session.scalars(
        db.select(AuditLog)
        .where(
            AuditLog.wedding_id == wedding.id,
            AuditLog.entity_type == "vendor",
            AuditLog.entity_id == str(vendor.id),
        )
        .order_by(AuditLog.created_at.desc())
        .limit(10)
    ).all()
    return render_template(
        "vendors/detail.html",
        vendor=vendor,
        tasks=tasks,
        documents=documents,
        activity=activity,
        category_labels=CATEGORY_LABELS,
        status_labels=STATUS_LABELS,
    )


@vendors_bp.route("/new", methods=["GET", "POST"])
@login_required
def create():
    wedding = current_wedding()
    form = VendorForm()
    if form.validate_on_submit():
        vendor = Vendor(wedding_id=wedding.id)
        apply_form(vendor, form)
        db.session.add(vendor)
        db.session.flush()
        sync_vendor_to_budget(vendor)
        audit(wedding.id, vendor, "create", f"נוסף הספק {vendor.name}")
        db.session.commit()
        flash("הספק נוסף.", "success")
        return redirect(url_for("vendors.index"))
    return render_template("vendors/form.html", form=form, title="הוספת ספק")


@vendors_bp.route("/<int:vendor_id>/edit", methods=["GET", "POST"])
@login_required
def edit(vendor_id: int):
    wedding = current_wedding()
    vendor = db.get_or_404(Vendor, vendor_id)
    if vendor.wedding_id != wedding.id or vendor.is_deleted:
        abort(404)
    form = VendorForm(obj=vendor)
    if form.validate_on_submit():
        apply_form(vendor, form)
        sync_vendor_to_budget(vendor)
        audit(wedding.id, vendor, "update", f"עודכן הספק {vendor.name}")
        db.session.commit()
        flash("הספק עודכן.", "success")
        return redirect(url_for("vendors.index"))
    return render_template(
        "vendors/form.html", form=form, title=f"עריכת {vendor.name}", vendor=vendor
    )


@vendors_bp.post("/<int:vendor_id>/favorite")
@login_required
def favorite(vendor_id: int):
    wedding = current_wedding()
    vendor = db.get_or_404(Vendor, vendor_id)
    if vendor.wedding_id != wedding.id or vendor.is_deleted:
        abort(404)
    vendor.is_favorite = not vendor.is_favorite
    audit(wedding.id, vendor, "favorite", f"עודכן סימון מועדף עבור {vendor.name}")
    db.session.commit()
    flash("סימון הספק עודכן.", "success")
    return redirect(request.referrer or url_for("vendors.index"))


@vendors_bp.post("/<int:vendor_id>/mark-paid")
@login_required
def mark_paid(vendor_id: int):
    wedding = current_wedding()
    vendor = db.get_or_404(Vendor, vendor_id)
    if vendor.wedding_id != wedding.id or vendor.is_deleted:
        abort(404)
    vendor.paid_amount = vendor.agreed_amount or 0
    sync_vendor_to_budget(vendor)
    audit(wedding.id, vendor, "paid", f"הספק {vendor.name} סומן כשולם במלואו")
    db.session.commit()
    flash("הספק סומן כשולם.", "success")
    return redirect(request.referrer or url_for("vendors.index"))


@vendors_bp.post("/<int:vendor_id>/delete")
@login_required
def delete(vendor_id: int):
    wedding = current_wedding()
    vendor = db.get_or_404(Vendor, vendor_id)
    if vendor.wedding_id != wedding.id:
        abort(404)
    vendor.soft_delete()
    sync_vendor_to_budget(vendor)
    audit(wedding.id, vendor, "delete", f"הספק {vendor.name} הועבר לסל המחזור")
    db.session.commit()
    flash("הספק הועבר לסל המחזור.", "success")
    return redirect(url_for("vendors.index"))


@vendors_bp.get("/export.xlsx")
@login_required
def export_excel():
    wedding = current_wedding()
    vendors = db.session.scalars(
        db.select(Vendor)
        .where(Vendor.wedding_id == wedding.id, Vendor.deleted_at.is_(None))
        .order_by(Vendor.category, Vendor.name)
    ).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "ספקים"
    ws.sheet_view.rightToLeft = True
    headers = [
        "ספק",
        "קטגוריה",
        "סטטוס",
        "איש קשר",
        "טלפון",
        "אימייל",
        "כתובת",
        "סכום שסוכם",
        "שולם",
        "יתרה",
        "חוזה",
        "תשלום הבא",
        "שעת הגעה",
        "דירוג",
        "מועדף",
        "אתר",
        "הערות",
    ]
    ws.append(headers)
    style_header(ws)
    for vendor in vendors:
        ws.append(
            [
                vendor.name,
                CATEGORY_LABELS.get(vendor.category, vendor.category),
                STATUS_LABELS.get(vendor.status, vendor.status),
                vendor.contact_name or "",
                vendor.phone or "",
                vendor.email or "",
                vendor.address or "",
                float(vendor.agreed_amount or 0),
                float(vendor.paid_amount or 0),
                float(vendor.balance),
                "כן" if vendor.contract_signed else "לא",
                vendor.next_payment_date.isoformat() if vendor.next_payment_date else "",
                vendor.arrival_time.strftime("%H:%M") if vendor.arrival_time else "",
                vendor.rating,
                "כן" if vendor.is_favorite else "לא",
                vendor.website_url or "",
                vendor.notes or "",
            ]
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [24, 17, 16, 18, 15, 25, 28, 15, 15, 15, 10, 14, 12, 10, 10, 35, 36]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    summary = wb.create_sheet("סיכום")
    summary.sheet_view.rightToLeft = True
    summary.append(["מדד", "ערך"])
    style_header(summary)
    summary.append(["מספר ספקים", len(vendors)])
    summary.append(
        ["ספקים שנסגרו", sum(1 for vendor in vendors if vendor.status in {"booked", "completed"})]
    )
    summary.append(
        [
            "חוזים חסרים",
            sum(
                1 for vendor in vendors if vendor.status == "booked" and not vendor.contract_signed
            ),
        ]
    )
    summary.append(["סה״כ שסוכם", sum(float(vendor.agreed_amount or 0) for vendor in vendors)])
    summary.append(["סה״כ שולם", sum(float(vendor.paid_amount or 0) for vendor in vendors)])
    summary.append(["סה״כ יתרה", sum(float(vendor.balance) for vendor in vendors)])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="wedding-vendors.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@vendors_bp.get("/share")
@login_required
def share_text():
    wedding = current_wedding()
    vendors = db.session.scalars(
        db.select(Vendor)
        .where(
            Vendor.wedding_id == wedding.id,
            Vendor.deleted_at.is_(None),
            Vendor.status != "cancelled",
        )
        .order_by(Vendor.category, Vendor.name)
    ).all()
    lines = ["🤝 רשימת ספקים", ""]
    for vendor in vendors:
        details = [f"*{vendor.name}* — {CATEGORY_LABELS.get(vendor.category, vendor.category)}"]
        if vendor.contact_name:
            details.append(f"איש קשר: {vendor.contact_name}")
        if vendor.phone:
            details.append(f"טלפון: {vendor.phone}")
        if vendor.arrival_time:
            details.append(f"הגעה: {vendor.arrival_time.strftime('%H:%M')}")
        if vendor.balance:
            details.append(f"יתרה: ₪{float(vendor.balance):,.0f}")
        lines.extend(details + [""])
    if not vendors:
        lines.append("אין ספקים פעילים.")
    text = "\n".join(lines)
    return render_template(
        "vendors/share.html", text=text, whatsapp_url=f"https://wa.me/?text={quote(text)}"
    )


def apply_form(vendor: Vendor, form: VendorForm) -> None:
    vendor.name = form.name.data.strip()
    vendor.category = form.category.data
    vendor.status = form.status.data
    vendor.contact_name = (form.contact_name.data or "").strip() or None
    vendor.phone = (form.phone.data or "").strip() or None
    vendor.email = (form.email.data or "").strip().lower() or None
    vendor.website_url = (form.website_url.data or "").strip() or None
    vendor.address = (form.address.data or "").strip() or None
    vendor.agreed_amount = form.agreed_amount.data or 0
    vendor.paid_amount = form.paid_amount.data or 0
    vendor.next_payment_date = form.next_payment_date.data
    vendor.arrival_time = form.arrival_time.data
    vendor.rating = form.rating.data or 0
    vendor.contract_signed = bool(form.contract_signed.data)
    vendor.is_favorite = bool(form.is_favorite.data)
    vendor.notes = (form.notes.data or "").strip() or None


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="40513B")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
