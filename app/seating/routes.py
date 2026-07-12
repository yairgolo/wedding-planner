from __future__ import annotations

from io import BytesIO

from flask import (
    Blueprint,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import or_

from app.extensions import db
from app.models import AuditLog, Guest, SeatingAssignment, SeatingTable, Wedding

from .forms import SeatingTableForm

seating_bp = Blueprint("seating", __name__, url_prefix="/seating")


def current_wedding() -> Wedding:
    wedding = db.session.scalar(db.select(Wedding).order_by(Wedding.id).limit(1))
    if not wedding:
        abort(404)
    return wedding


def expected_seats(guest: Guest) -> int:
    if guest.rsvp_status == "declined":
        return 0
    if guest.rsvp_status == "confirmed":
        return max(guest.confirmed_count, 1)
    return max(guest.invited_count, 1)


def add_audit(wedding_id: int, entity_type: str, entity_id: str, action: str, text: str) -> None:
    db.session.add(
        AuditLog(
            wedding_id=wedding_id,
            user_id=current_user.id,
            entity_type=entity_type,
            entity_id=entity_id,
            action=action,
            description=text,
        )
    )


@seating_bp.get("")
@login_required
def index():
    wedding = current_wedding()
    q = request.args.get("q", "").strip()
    tables = db.session.scalars(
        db.select(SeatingTable)
        .where(SeatingTable.wedding_id == wedding.id)
        .order_by(SeatingTable.number)
    ).all()

    guest_stmt = (
        db.select(Guest)
        .where(
            Guest.wedding_id == wedding.id,
            Guest.deleted_at.is_(None),
            Guest.rsvp_status != "declined",
        )
        .order_by(Guest.last_name, Guest.first_name)
    )
    if q:
        pattern = f"%{q}%"
        guest_stmt = guest_stmt.where(
            or_(
                Guest.first_name.ilike(pattern),
                Guest.last_name.ilike(pattern),
                Guest.phone.ilike(pattern),
                Guest.group_name.ilike(pattern),
            )
        )
    guests = db.session.scalars(guest_stmt).all()
    assignments = db.session.scalars(
        db.select(SeatingAssignment).where(SeatingAssignment.wedding_id == wedding.id)
    ).all()
    assignment_by_guest = {assignment.guest_id: assignment for assignment in assignments}
    unseated = [guest for guest in guests if guest.id not in assignment_by_guest]

    stats = {
        "tables": len(tables),
        "capacity": sum(table.capacity for table in tables),
        "seated": sum(assignment.seat_count for assignment in assignments),
        "unseated": sum(expected_seats(guest) for guest in unseated),
    }
    return render_template(
        "seating/index.html",
        wedding=wedding,
        tables=tables,
        unseated=unseated,
        stats=stats,
        expected_seats=expected_seats,
        q=q,
    )


@seating_bp.route("/tables/new", methods=["GET", "POST"])
@login_required
def create_table():
    wedding = current_wedding()
    form = SeatingTableForm()
    if form.validate_on_submit():
        duplicate = db.session.scalar(
            db.select(SeatingTable).where(
                SeatingTable.wedding_id == wedding.id,
                SeatingTable.number == form.number.data.strip(),
            )
        )
        if duplicate:
            form.number.errors.append("כבר קיים שולחן עם המספר הזה.")
        else:
            table = SeatingTable(wedding_id=wedding.id)
            apply_table_form(table, form)
            db.session.add(table)
            db.session.flush()
            add_audit(
                wedding.id,
                "seating_table",
                str(table.id),
                "create",
                f"נוצר {table.display_name}",
            )
            db.session.commit()
            flash("השולחן נוסף בהצלחה.", "success")
            return redirect(url_for("seating.index"))
    return render_template("seating/table_form.html", form=form, title="הוספת שולחן")


@seating_bp.route("/tables/<int:table_id>/edit", methods=["GET", "POST"])
@login_required
def edit_table(table_id: int):
    wedding = current_wedding()
    table = db.get_or_404(SeatingTable, table_id)
    if table.wedding_id != wedding.id:
        abort(404)
    form = SeatingTableForm(obj=table)
    if form.validate_on_submit():
        new_capacity = form.capacity.data
        if new_capacity < table.occupied_seats:
            form.capacity.errors.append(
                f"יש כרגע {table.occupied_seats} מקומות משובצים. לא ניתן להקטין מתחת לכמות זו."
            )
        else:
            duplicate = db.session.scalar(
                db.select(SeatingTable).where(
                    SeatingTable.wedding_id == wedding.id,
                    SeatingTable.number == form.number.data.strip(),
                    SeatingTable.id != table.id,
                )
            )
            if duplicate:
                form.number.errors.append("כבר קיים שולחן עם המספר הזה.")
            else:
                apply_table_form(table, form)
                add_audit(
                    wedding.id,
                    "seating_table",
                    str(table.id),
                    "update",
                    f"עודכן {table.display_name}",
                )
                db.session.commit()
                flash("השולחן עודכן.", "success")
                return redirect(url_for("seating.index"))
    return render_template(
        "seating/table_form.html",
        form=form,
        title=f"עריכת {table.display_name}",
    )


@seating_bp.post("/tables/<int:table_id>/delete")
@login_required
def delete_table(table_id: int):
    wedding = current_wedding()
    table = db.get_or_404(SeatingTable, table_id)
    if table.wedding_id != wedding.id:
        abort(404)
    if table.assignments:
        flash("לא ניתן למחוק שולחן שיש בו מוזמנים. יש להעביר או להסיר אותם קודם.", "danger")
    else:
        name = table.display_name
        db.session.delete(table)
        add_audit(wedding.id, "seating_table", str(table.id), "delete", f"נמחק {name}")
        db.session.commit()
        flash("השולחן נמחק.", "success")
    return redirect(url_for("seating.index"))


@seating_bp.post("/assign")
@login_required
def assign():
    wedding = current_wedding()
    payload = request.get_json(silent=True) or request.form
    try:
        guest_id = int(payload.get("guest_id", 0))
        table_id = int(payload.get("table_id", 0))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "בקשה לא תקינה."}), 400

    guest = db.session.get(Guest, guest_id)
    table = db.session.get(SeatingTable, table_id)
    if not guest or not table or guest.wedding_id != wedding.id or table.wedding_id != wedding.id:
        return jsonify({"ok": False, "error": "המוזמן או השולחן לא נמצאו."}), 404
    if guest.is_deleted or guest.rsvp_status == "declined":
        return jsonify({"ok": False, "error": "לא ניתן לשבץ מוזמן זה."}), 400

    seats = expected_seats(guest)
    assignment = db.session.scalar(
        db.select(SeatingAssignment).where(
            SeatingAssignment.wedding_id == wedding.id,
            SeatingAssignment.guest_id == guest.id,
        )
    )
    previous_seats = assignment.seat_count if assignment and assignment.table_id == table.id else 0
    projected = table.occupied_seats - previous_seats + seats
    if projected > table.capacity:
        return jsonify(
            {
                "ok": False,
                "error": (
                    f"אין מספיק מקום. נדרשים {seats} מקומות "
                    f"ונשארו {table.available_seats + previous_seats}."
                ),
            }
        ), 409

    if not assignment:
        assignment = SeatingAssignment(
            wedding_id=wedding.id, guest_id=guest.id, table_id=table.id, seat_count=seats
        )
        db.session.add(assignment)
    else:
        assignment.table_id = table.id
        assignment.seat_count = seats
    guest.table_number = table.number
    add_audit(
        wedding.id,
        "guest",
        str(guest.id),
        "seat_assign",
        f"{guest.full_name} שובץ ל{table.display_name}",
    )
    db.session.commit()
    return jsonify({"ok": True, "message": f"{guest.full_name} שובץ ל{table.display_name}."})


@seating_bp.post("/unassign")
@login_required
def unassign():
    wedding = current_wedding()
    payload = request.get_json(silent=True) or request.form
    try:
        guest_id = int(payload.get("guest_id", 0))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "בקשה לא תקינה."}), 400
    guest = db.session.get(Guest, guest_id)
    if not guest or guest.wedding_id != wedding.id:
        return jsonify({"ok": False, "error": "המוזמן לא נמצא."}), 404
    assignment = db.session.scalar(
        db.select(SeatingAssignment).where(
            SeatingAssignment.wedding_id == wedding.id,
            SeatingAssignment.guest_id == guest.id,
        )
    )
    if assignment:
        db.session.delete(assignment)
    guest.table_number = None
    add_audit(
        wedding.id,
        "guest",
        str(guest.id),
        "seat_unassign",
        f"הוסר השיבוץ של {guest.full_name}",
    )
    db.session.commit()
    return jsonify({"ok": True, "message": f"השיבוץ של {guest.full_name} הוסר."})


@seating_bp.get("/event-day")
@login_required
def event_day():
    wedding = current_wedding()
    q = request.args.get("q", "").strip()
    results = []
    if q:
        pattern = f"%{q}%"
        results = db.session.scalars(
            db.select(Guest)
            .where(
                Guest.wedding_id == wedding.id,
                Guest.deleted_at.is_(None),
                or_(
                    Guest.first_name.ilike(pattern),
                    Guest.last_name.ilike(pattern),
                    Guest.phone.ilike(pattern),
                ),
            )
            .order_by(Guest.last_name, Guest.first_name)
            .limit(25)
        ).all()
    return render_template("seating/event_day.html", wedding=wedding, results=results, q=q)


@seating_bp.get("/export.xlsx")
@login_required
def export_excel():
    wedding = current_wedding()
    tables = db.session.scalars(
        db.select(SeatingTable)
        .where(SeatingTable.wedding_id == wedding.id)
        .order_by(SeatingTable.number)
    ).all()
    unseated = db.session.scalars(
        db.select(Guest)
        .outerjoin(SeatingAssignment, SeatingAssignment.guest_id == Guest.id)
        .where(
            Guest.wedding_id == wedding.id,
            Guest.deleted_at.is_(None),
            Guest.rsvp_status != "declined",
            SeatingAssignment.id.is_(None),
        )
        .order_by(Guest.last_name, Guest.first_name)
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "הושבה"
    ws.sheet_view.rightToLeft = True
    headers = ["שולחן", "שם שולחן", "אזור", "שם מוזמן", "טלפון", "כמות", "צד", "תזונה", "הערות"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="40513B")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    side_labels = {"groom": "חתן", "bride": "כלה", "shared": "משותף"}
    for table in tables:
        for assignment in sorted(table.assignments, key=lambda item: item.guest.full_name):
            guest = assignment.guest
            ws.append([
                table.number,
                table.name or "",
                table.zone or "",
                guest.full_name,
                guest.phone or "",
                assignment.seat_count,
                side_labels.get(guest.side, guest.side),
                guest.diet_notes or guest.diet,
                guest.notes or "",
            ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, width in enumerate([12, 20, 16, 28, 17, 10, 12, 22, 34], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    pending = wb.create_sheet("ללא שולחן")
    pending.sheet_view.rightToLeft = True
    pending.append(["שם", "טלפון", "כמות צפויה", "סטטוס", "קבוצה"])
    for cell in pending[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    for guest in unseated:
        pending.append([
            guest.full_name,
            guest.phone or "",
            expected_seats(guest),
            guest.rsvp_status,
            guest.group_name or "",
        ])
    summary = wb.create_sheet("סיכום")
    summary.sheet_view.rightToLeft = True
    summary.append(["מדד", "כמות"])
    summary.append(["שולחנות", len(tables)])
    summary.append(["סה״כ מקומות", sum(table.capacity for table in tables)])
    summary.append(["משובצים", sum(table.occupied_seats for table in tables)])
    summary.append(["ללא שולחן", sum(expected_seats(guest) for guest in unseated)])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="wedding-seating.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def apply_table_form(table: SeatingTable, form: SeatingTableForm) -> None:
    table.number = form.number.data.strip()
    table.name = (form.name.data or "").strip() or None
    table.capacity = form.capacity.data
    table.shape = form.shape.data
    table.zone = (form.zone.data or "").strip() or None
    table.notes = (form.notes.data or "").strip() or None
