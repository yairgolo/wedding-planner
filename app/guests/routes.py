from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO

from flask import Blueprint, abort, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import or_

from app.extensions import db
from app.models import AuditLog, Family, Guest, Wedding
from app.services.guest_roundtrip import (
    DIET_TO_LABEL,
    EXPORT_SCHEMA_VERSION,
    SIDE_TO_LABEL,
    STATUS_TO_LABEL,
    guest_to_row,
)
from app.services.guest_roundtrip import (
    HEADERS as ROUNDTRIP_HEADERS,
)

from .forms import FamilyForm, GuestForm, RSVPForm

guests_bp = Blueprint("guests", __name__, url_prefix="/guests")
rsvp_bp = Blueprint("rsvp", __name__, url_prefix="/rsvp")

STATUS_LABELS = {"pending": "ממתין", "confirmed": "מגיע", "declined": "לא מגיע", "maybe": "אולי"}
SIDE_LABELS = {"groom": "חתן", "bride": "כלה", "shared": "משותף"}
DIET_LABELS = {
    "regular": "רגיל",
    "vegetarian": "צמחוני",
    "vegan": "טבעוני",
    "gluten_free": "ללא גלוטן",
    "child": "מנת ילדים",
    "other": "אחר",
}


def current_wedding() -> Wedding:
    wedding = db.session.scalar(db.select(Wedding).order_by(Wedding.id).limit(1))
    if not wedding:
        abort(404)
    return wedding


def family_choices(wedding_id: int):
    families = db.session.scalars(
        db.select(Family).where(Family.wedding_id == wedding_id).order_by(Family.name)
    ).all()
    return [(0, "ללא משפחה")] + [(family.id, family.name) for family in families]


def audit(wedding_id: int, entity: Guest | Family, action: str, description: str) -> None:
    db.session.add(
        AuditLog(
            wedding_id=wedding_id,
            user_id=current_user.id if current_user.is_authenticated else None,
            entity_type=entity.__class__.__name__.lower(),
            entity_id=str(entity.id or "new"),
            action=action,
            description=description,
        )
    )


@guests_bp.get("")
@login_required
def index():
    wedding = current_wedding()
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip()
    side = request.args.get("side", "").strip()
    vip = request.args.get("vip", "").strip()
    sent = request.args.get("sent", "").strip()
    view = request.args.get("view", "list").strip()

    stmt = db.select(Guest).where(Guest.wedding_id == wedding.id, Guest.deleted_at.is_(None))
    if q:
        pattern = f"%{q}%"
        stmt = stmt.outerjoin(Family).where(
            or_(
                Guest.first_name.ilike(pattern),
                Guest.last_name.ilike(pattern),
                Guest.phone.ilike(pattern),
                Guest.email.ilike(pattern),
                Guest.group_name.ilike(pattern),
                Family.name.ilike(pattern),
            )
        )
    if status:
        stmt = stmt.where(Guest.rsvp_status == status)
    if side:
        stmt = stmt.where(Guest.side == side)
    if vip == "1":
        stmt = stmt.where(Guest.is_vip.is_(True))
    if sent == "yes":
        stmt = stmt.where(Guest.invitation_sent.is_(True))
    elif sent == "no":
        stmt = stmt.where(Guest.invitation_sent.is_(False))

    guests = db.session.scalars(stmt.order_by(Guest.created_at.desc())).all()
    all_active = db.session.scalars(
        db.select(Guest).where(Guest.wedding_id == wedding.id, Guest.deleted_at.is_(None))
    ).all()
    stats = {
        "records": len(all_active),
        "invited": sum(g.invited_count for g in all_active),
        "confirmed": sum(g.confirmed_count for g in all_active if g.rsvp_status == "confirmed"),
        "pending": sum(
            g.invited_count for g in all_active if g.rsvp_status in {"pending", "maybe"}
        ),
        "declined": sum(g.invited_count for g in all_active if g.rsvp_status == "declined"),
        "unsent": sum(1 for g in all_active if not g.invitation_sent),
    }
    family_groups = []
    if view == "family":
        grouped = {}
        singles = []
        for guest in guests:
            key = guest.family.name if guest.family else (guest.group_name or "")
            if key:
                grouped.setdefault(key, []).append(guest)
            else:
                singles.append(guest)
        for name, members in sorted(grouped.items(), key=lambda pair: pair[0]):
            family_groups.append({
                "name": name,
                "members": members,
                "invited": sum(member.invited_count for member in members),
                "confirmed": sum(
                    member.confirmed_count
                    for member in members
                    if member.rsvp_status == "confirmed"
                ),
                "pending": sum(
                    member.invited_count
                    for member in members
                    if member.rsvp_status in {"pending", "maybe"}
                ),
                "tables": sorted(
                    {member.table_number for member in members if member.table_number}
                ),
            })
        if singles:
            family_groups.append({
                "name": "מוזמנים ללא משפחה", "members": singles,
                "invited": sum(member.invited_count for member in singles),
                "confirmed": sum(
                    member.confirmed_count
                    for member in singles
                    if member.rsvp_status == "confirmed"
                ),
                "pending": sum(
                    member.invited_count
                    for member in singles
                    if member.rsvp_status in {"pending", "maybe"}
                ),
                "tables": sorted(
                    {member.table_number for member in singles if member.table_number}
                ),
            })

    return render_template(
        "guests/index.html",
        wedding=wedding,
        guests=guests,
        stats=stats,
        status_labels=STATUS_LABELS,
        side_labels=SIDE_LABELS,
        q=q,
        view=view,
        family_groups=family_groups,
    )


@guests_bp.route("/new", methods=["GET", "POST"])
@login_required
def create():
    wedding = current_wedding()
    form = GuestForm()
    form.family_id.choices = family_choices(wedding.id)
    if form.validate_on_submit():
        guest = Guest(wedding_id=wedding.id)
        apply_form(guest, form)
        db.session.add(guest)
        db.session.flush()
        audit(wedding.id, guest, "create", f"נוסף המוזמן {guest.full_name}")
        db.session.commit()
        flash("המוזמן נוסף בהצלחה.", "success")
        return redirect(url_for("guests.index"))
    return render_template("guests/form.html", form=form, title="הוספת מוזמן")


@guests_bp.route("/<int:guest_id>/edit", methods=["GET", "POST"])
@login_required
def edit(guest_id: int):
    wedding = current_wedding()
    guest = db.get_or_404(Guest, guest_id)
    if guest.wedding_id != wedding.id or guest.is_deleted:
        abort(404)
    form = GuestForm(obj=guest)
    form.family_id.choices = family_choices(wedding.id)
    if request.method == "GET":
        form.family_id.data = guest.family_id or 0
    if form.validate_on_submit():
        apply_form(guest, form)
        audit(wedding.id, guest, "update", f"עודכן המוזמן {guest.full_name}")
        db.session.commit()
        flash("פרטי המוזמן עודכנו.", "success")
        return redirect(url_for("guests.index"))
    return render_template(
        "guests/form.html", form=form, title=f"עריכת {guest.full_name}", guest=guest
    )


@guests_bp.post("/<int:guest_id>/delete")
@login_required
def delete(guest_id: int):
    wedding = current_wedding()
    guest = db.get_or_404(Guest, guest_id)
    if guest.wedding_id != wedding.id:
        abort(404)
    guest.soft_delete()
    audit(wedding.id, guest, "delete", f"המוזמן {guest.full_name} הועבר לסל המחזור")
    db.session.commit()
    flash("המוזמן הועבר לסל המחזור.", "success")
    return redirect(url_for("guests.index"))


@guests_bp.post("/<int:guest_id>/sent")
@login_required
def mark_sent(guest_id: int):
    wedding = current_wedding()
    guest = db.get_or_404(Guest, guest_id)
    if guest.wedding_id != wedding.id or guest.is_deleted:
        abort(404)
    guest.invitation_sent = True
    guest.invitation_sent_at = datetime.now(UTC)
    guest.invitation_attempts += 1
    audit(wedding.id, guest, "invitation_sent", f"סומנה שליחת הזמנה ל{guest.full_name}")
    db.session.commit()
    flash("סומן שההזמנה נשלחה.", "success")
    return redirect(request.referrer or url_for("guests.index"))


@guests_bp.route("/families/new", methods=["GET", "POST"])
@login_required
def create_family():
    wedding = current_wedding()
    form = FamilyForm()
    if form.validate_on_submit():
        family = Family(wedding_id=wedding.id, name=form.name.data.strip(), notes=form.notes.data)
        db.session.add(family)
        db.session.flush()
        audit(wedding.id, family, "create", f"נוצרה המשפחה {family.name}")
        db.session.commit()
        flash("המשפחה נוספה.", "success")
        return redirect(url_for("guests.create"))
    return render_template("guests/family_form.html", form=form)


@guests_bp.get("/export.xlsx")
@login_required
def export_excel():
    """Export a workbook that can be edited and imported back without data loss."""
    wedding = current_wedding()
    guests = db.session.scalars(
        db.select(Guest)
        .where(Guest.wedding_id == wedding.id, Guest.deleted_at.is_(None))
        .order_by(Guest.last_name, Guest.first_name)
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "מוזמנים"
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.append(ROUNDTRIP_HEADERS)

    header_fill = PatternFill("solid", fgColor="40513B")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for guest in guests:
        ws.append(guest_to_row(guest))

    # System identifiers are visually separated. They remain editable so users can
    # freely append new rows in Excel; the importer never trusts them across weddings.
    system_fill = PatternFill("solid", fgColor="E7E2D8")
    for row in ws.iter_rows(min_row=2):
        row[0].fill = system_fill
        row[1].fill = system_fill

    widths = [14, 38, 18, 18, 17, 28, 12, 18, 20, 10, 10, 14, 8, 18, 24, 12, 16, 36]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 28
    # Friendly dropdown values keep edited workbooks compatible with import.
    validations = [
        ("G", list(SIDE_TO_LABEL.values())),
        ("L", list(STATUS_TO_LABEL.values())),
        ("M", ["כן", "לא"]),
        ("N", list(DIET_TO_LABEL.values())),
        ("Q", ["כן", "לא"]),
    ]
    max_row = max(ws.max_row + 1000, 1001)
    for column, options in validations:
        formula = '"' + ",".join(options) + '"'
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "יש לבחור ערך מהרשימה"
        validation.errorTitle = "ערך לא מוכר"
        ws.add_data_validation(validation)
        validation.add(f"{column}2:{column}{max_row}")

    summary = wb.create_sheet("סיכום")
    summary.sheet_view.rightToLeft = True
    summary.append(["מדד", "כמות"])
    summary.append(["רשומות", len(guests)])
    summary.append(["סה״כ מוזמנים", sum(g.invited_count for g in guests)])
    summary.append(
        ["סה״כ מאשרים", sum(g.confirmed_count for g in guests if g.rsvp_status == "confirmed")]
    )
    for cell in summary[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    summary.column_dimensions["A"].width = 25
    summary.column_dimensions["B"].width = 15

    meta = wb.create_sheet("_מערכת")
    meta.sheet_state = "hidden"
    meta.append(["schema", EXPORT_SCHEMA_VERSION])
    meta.append(["wedding_id", wedding.id])
    meta.append(["instructions", "Do not delete the UUID column for existing guests."])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="wedding-guests-roundtrip.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def apply_form(guest: Guest, form: GuestForm) -> None:
    guest.first_name = form.first_name.data.strip()
    guest.last_name = (form.last_name.data or "").strip() or None
    guest.phone = (form.phone.data or "").strip() or None
    guest.email = (form.email.data or "").strip().lower() or None
    guest.side = form.side.data
    guest.group_name = (form.group_name.data or "").strip() or None
    guest.family_id = form.family_id.data or None
    guest.invited_count = form.invited_count.data
    guest.confirmed_count = form.confirmed_count.data
    guest.rsvp_status = form.rsvp_status.data
    guest.is_vip = bool(form.is_vip.data)
    guest.diet = form.diet.data
    guest.diet_notes = (form.diet_notes.data or "").strip() or None
    guest.table_number = (form.table_number.data or "").strip() or None
    guest.notes = (form.notes.data or "").strip() or None


@rsvp_bp.route("/<token>", methods=["GET", "POST"])
def respond(token: str):
    guest = db.session.scalar(
        db.select(Guest).where(Guest.rsvp_token == token, Guest.deleted_at.is_(None))
    )
    if not guest:
        abort(404)
    form = RSVPForm()
    if request.method == "GET":
        form.status.data = guest.rsvp_status if guest.rsvp_status != "pending" else "confirmed"
        form.confirmed_count.data = guest.confirmed_count or min(guest.invited_count, 1)
        form.diet_notes.data = guest.diet_notes
        form.message.data = guest.rsvp_message
    if form.validate_on_submit():
        if form.confirmed_count.data > guest.invited_count:
            form.confirmed_count.errors.append("הכמות לא יכולה להיות גדולה ממספר המוזמנים בהזמנה.")
        else:
            guest.rsvp_status = form.status.data
            guest.confirmed_count = (
                form.confirmed_count.data if form.status.data == "confirmed" else 0
            )
            guest.diet_notes = (form.diet_notes.data or "").strip() or None
            guest.rsvp_message = (form.message.data or "").strip() or None
            guest.rsvp_updated_at = datetime.now(UTC)
            db.session.add(
                AuditLog(
                    wedding_id=guest.wedding_id,
                    user_id=None,
                    entity_type="guest",
                    entity_id=str(guest.id),
                    action="rsvp",
                    description=f"{guest.full_name} עדכן אישור הגעה",
                )
            )
            db.session.commit()
            return render_template("rsvp/success.html", guest=guest)
    return render_template("rsvp/form.html", guest=guest, wedding=guest.wedding, form=form)
