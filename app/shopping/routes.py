from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from urllib.parse import quote

from flask import Blueprint, abort, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import or_

from app.extensions import db
from app.models import AuditLog, ShoppingItem, Wedding

from .forms import ShoppingItemForm

shopping_bp = Blueprint("shopping", __name__, url_prefix="/shopping")

CATEGORY_LABELS = {
    "wedding": "חתונה",
    "home": "בית",
    "clothing": "בגדים",
    "gifts": "מתנות",
    "general": "כללי",
    "other": "אחר",
}
STATUS_LABELS = {
    "planned": "מתוכנן",
    "ordered": "הוזמן",
    "purchased": "נקנה",
    "cancelled": "בוטל",
}
PRIORITY_LABELS = {"low": "נמוכה", "medium": "רגילה", "high": "גבוהה", "urgent": "דחופה"}


def current_wedding() -> Wedding:
    wedding = db.session.scalar(db.select(Wedding).order_by(Wedding.id).limit(1))
    if not wedding:
        abort(404)
    return wedding


def audit(wedding_id: int, item: ShoppingItem, action: str, description: str) -> None:
    db.session.add(
        AuditLog(
            wedding_id=wedding_id,
            user_id=current_user.id,
            entity_type="shopping_item",
            entity_id=str(item.id or "new"),
            action=action,
            description=description,
        )
    )


@shopping_bp.get("")
@login_required
def index():
    wedding = current_wedding()
    q = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    status = request.args.get("status", "").strip()
    wishlist = request.args.get("wishlist", "").strip()

    stmt = db.select(ShoppingItem).where(
        ShoppingItem.wedding_id == wedding.id, ShoppingItem.deleted_at.is_(None)
    )
    if q:
        pattern = f"%{q}%"
        stmt = stmt.where(
            or_(
                ShoppingItem.name.ilike(pattern),
                ShoppingItem.store_name.ilike(pattern),
                ShoppingItem.notes.ilike(pattern),
            )
        )
    if category:
        stmt = stmt.where(ShoppingItem.category == category)
    if status:
        stmt = stmt.where(ShoppingItem.status == status)
    if wishlist == "1":
        stmt = stmt.where(ShoppingItem.is_wishlist.is_(True))

    items = db.session.scalars(
        stmt.order_by(
            ShoppingItem.is_wishlist,
            ShoppingItem.due_date,
            ShoppingItem.created_at.desc(),
        )
    ).all()
    all_items = db.session.scalars(
        db.select(ShoppingItem).where(
            ShoppingItem.wedding_id == wedding.id, ShoppingItem.deleted_at.is_(None)
        )
    ).all()
    stats = {
        "total": len(all_items),
        "purchased": sum(1 for item in all_items if item.is_purchased),
        "remaining": sum(1 for item in all_items if item.status not in {"purchased", "cancelled"}),
        "estimated": sum(
            float(item.total_estimated)
            for item in all_items
            if item.status != "cancelled"
        ),
        "actual": sum(float(item.total_actual) for item in all_items if item.status == "purchased"),
    }
    return render_template(
        "shopping/index.html",
        wedding=wedding,
        items=items,
        stats=stats,
        category_labels=CATEGORY_LABELS,
        status_labels=STATUS_LABELS,
        priority_labels=PRIORITY_LABELS,
        q=q,
    )


@shopping_bp.route("/new", methods=["GET", "POST"])
@login_required
def create():
    wedding = current_wedding()
    form = ShoppingItemForm()
    if form.validate_on_submit():
        item = ShoppingItem(wedding_id=wedding.id)
        apply_form(item, form)
        db.session.add(item)
        db.session.flush()
        audit(wedding.id, item, "create", f"נוסף פריט הקנייה {item.name}")
        db.session.commit()
        flash("הפריט נוסף לרשימת הקניות.", "success")
        return redirect(url_for("shopping.index"))
    return render_template("shopping/form.html", form=form, title="הוספת פריט")


@shopping_bp.route("/<int:item_id>/edit", methods=["GET", "POST"])
@login_required
def edit(item_id: int):
    wedding = current_wedding()
    item = db.get_or_404(ShoppingItem, item_id)
    if item.wedding_id != wedding.id or item.is_deleted:
        abort(404)
    form = ShoppingItemForm(obj=item)
    if form.validate_on_submit():
        previous_status = item.status
        apply_form(item, form)
        if previous_status != "purchased" and item.status == "purchased":
            item.purchased_at = datetime.now(UTC)
        audit(wedding.id, item, "update", f"עודכן פריט הקנייה {item.name}")
        db.session.commit()
        flash("הפריט עודכן.", "success")
        return redirect(url_for("shopping.index"))
    return render_template("shopping/form.html", form=form, title=f"עריכת {item.name}", item=item)


@shopping_bp.post("/<int:item_id>/toggle")
@login_required
def toggle(item_id: int):
    wedding = current_wedding()
    item = db.get_or_404(ShoppingItem, item_id)
    if item.wedding_id != wedding.id or item.is_deleted:
        abort(404)
    if item.is_purchased:
        item.status = "planned"
        item.purchased_at = None
        action = "reopen"
        message = "הפריט הוחזר לרשימה."
    else:
        item.mark_purchased()
        action = "purchased"
        message = "הפריט סומן כנקנה."
    audit(wedding.id, item, action, f"עודכן סטטוס הקנייה {item.name}")
    db.session.commit()
    flash(message, "success")
    return redirect(request.referrer or url_for("shopping.index"))


@shopping_bp.post("/<int:item_id>/delete")
@login_required
def delete(item_id: int):
    wedding = current_wedding()
    item = db.get_or_404(ShoppingItem, item_id)
    if item.wedding_id != wedding.id:
        abort(404)
    item.soft_delete()
    audit(wedding.id, item, "delete", f"פריט הקנייה {item.name} הועבר לסל המחזור")
    db.session.commit()
    flash("הפריט הועבר לסל המחזור.", "success")
    return redirect(url_for("shopping.index"))


@shopping_bp.get("/export.xlsx")
@login_required
def export_excel():
    wedding = current_wedding()
    items = db.session.scalars(
        db.select(ShoppingItem)
        .where(ShoppingItem.wedding_id == wedding.id, ShoppingItem.deleted_at.is_(None))
        .order_by(ShoppingItem.category, ShoppingItem.name)
    ).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "קניות"
    ws.sheet_view.rightToLeft = True
    headers = [
        "פריט", "קטגוריה", "סטטוס", "עדיפות", "כמות", "מחיר משוער ליחידה",
        "מחיר בפועל ליחידה", "סה״כ משוער", "סה״כ בפועל", "חנות", "תאריך יעד",
        "Wishlist", "קישור", "הערות",
    ]
    ws.append(headers)
    style_header(ws)
    for item in items:
        ws.append([
            item.name,
            CATEGORY_LABELS.get(item.category, item.category),
            STATUS_LABELS.get(item.status, item.status),
            PRIORITY_LABELS.get(item.priority, item.priority),
            item.quantity,
            float(item.estimated_price or 0),
            float(item.actual_price or 0),
            float(item.total_estimated),
            float(item.total_actual),
            item.store_name or "",
            item.due_date.isoformat() if item.due_date else "",
            "כן" if item.is_wishlist else "לא",
            item.product_url or "",
            item.notes or "",
        ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [28, 14, 14, 12, 9, 18, 18, 16, 16, 20, 14, 12, 36, 32]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    summary = wb.create_sheet("סיכום")
    summary.sheet_view.rightToLeft = True
    summary.append(["מדד", "ערך"])
    style_header(summary)
    summary.append(["מספר פריטים", len(items)])
    summary.append(["נקנו", sum(1 for item in items if item.is_purchased)])
    summary.append(
        ["נותרו", sum(1 for item in items if item.status not in {"purchased", "cancelled"})]
    )
    summary.append(["עלות משוערת", sum(float(item.total_estimated) for item in items)])
    summary.append(
        ["עלות בפועל", sum(float(item.total_actual) for item in items if item.is_purchased)]
    )
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="wedding-shopping.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@shopping_bp.get("/share")
@login_required
def share_text():
    wedding = current_wedding()
    category = request.args.get("category", "").strip()
    stmt = db.select(ShoppingItem).where(
        ShoppingItem.wedding_id == wedding.id,
        ShoppingItem.deleted_at.is_(None),
        ShoppingItem.status.notin_(["purchased", "cancelled"]),
    )
    if category:
        stmt = stmt.where(ShoppingItem.category == category)
    items = db.session.scalars(stmt.order_by(ShoppingItem.category, ShoppingItem.name)).all()
    lines = ["🛒 רשימת קניות", ""]
    current_category = None
    for item in items:
        if item.category != current_category:
            current_category = item.category
            lines.extend([f"*{CATEGORY_LABELS.get(current_category, current_category)}*", ""])
        price = f" — ₪{float(item.total_estimated):,.0f}" if item.total_estimated else ""
        lines.append(f"☐ {item.name} × {item.quantity}{price}")
    if not items:
        lines.append("אין פריטים פתוחים 🎉")
    text = "\n".join(lines)
    return render_template("shopping/share.html", text=text, whatsapp_url=f"https://wa.me/?text={quote(text)}")


def apply_form(item: ShoppingItem, form: ShoppingItemForm) -> None:
    item.name = form.name.data.strip()
    item.category = form.category.data
    item.status = form.status.data
    item.priority = form.priority.data
    item.quantity = form.quantity.data
    item.estimated_price = form.estimated_price.data or 0
    item.actual_price = form.actual_price.data or 0
    item.store_name = (form.store_name.data or "").strip() or None
    item.product_url = (form.product_url.data or "").strip() or None
    item.due_date = form.due_date.data
    item.is_wishlist = bool(form.is_wishlist.data)
    item.notes = (form.notes.data or "").strip() or None


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="40513B")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
