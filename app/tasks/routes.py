from __future__ import annotations

from io import BytesIO
from urllib.parse import quote

from flask import Blueprint, abort, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import or_

from app.extensions import db
from app.models import AuditLog, Task, Vendor, Wedding

from .forms import TaskForm

tasks_bp = Blueprint("tasks", __name__, url_prefix="/tasks")

CATEGORY_LABELS = {
    "wedding": "חתונה",
    "home": "בית",
    "documents": "מסמכים",
    "payments": "תשלומים",
    "guests": "מוזמנים",
    "seating": "הושבה",
    "shopping": "קניות",
    "general": "כללי",
}
STATUS_LABELS = {"todo": "לביצוע", "doing": "בתהליך", "done": "הושלם"}
PRIORITY_LABELS = {"low": "נמוכה", "medium": "רגילה", "high": "גבוהה", "urgent": "דחופה"}


def current_wedding() -> Wedding:
    wedding = db.session.scalar(db.select(Wedding).order_by(Wedding.id).limit(1))
    if not wedding:
        abort(404)
    return wedding


def audit(wedding_id: int, task: Task, action: str, description: str) -> None:
    db.session.add(
        AuditLog(
            wedding_id=wedding_id,
            user_id=current_user.id,
            entity_type="task",
            entity_id=str(task.id or "new"),
            action=action,
            description=description,
        )
    )


def configure_vendor_choices(form: TaskForm, wedding_id: int) -> None:
    vendors = db.session.scalars(
        db.select(Vendor)
        .where(
            Vendor.wedding_id == wedding_id,
            Vendor.deleted_at.is_(None),
            Vendor.status != "cancelled",
        )
        .order_by(Vendor.name)
    ).all()
    form.related_vendor_id.choices = [(0, "ללא ספק קשור")] + [
        (vendor.id, vendor.name) for vendor in vendors
    ]


@tasks_bp.get("")
@login_required
def index():
    wedding = current_wedding()
    q = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    status = request.args.get("status", "").strip()
    priority = request.args.get("priority", "").strip()
    assignee = request.args.get("assignee", "").strip()

    stmt = db.select(Task).where(Task.wedding_id == wedding.id, Task.deleted_at.is_(None))
    if q:
        pattern = f"%{q}%"
        stmt = stmt.where(
            or_(
                Task.title.ilike(pattern),
                Task.notes.ilike(pattern),
                Task.assigned_to.ilike(pattern),
            )
        )
    if category:
        stmt = stmt.where(Task.category == category)
    if status:
        stmt = stmt.where(Task.status == status)
    if priority:
        stmt = stmt.where(Task.priority == priority)
    if assignee:
        stmt = stmt.where(Task.assigned_to == assignee)

    tasks = db.session.scalars(
        stmt.order_by(Task.status, Task.due_date.is_(None), Task.due_date, Task.created_at.desc())
    ).all()
    all_tasks = db.session.scalars(
        db.select(Task).where(Task.wedding_id == wedding.id, Task.deleted_at.is_(None))
    ).all()
    assignees = sorted({task.assigned_to for task in all_tasks if task.assigned_to})
    stats = {
        "total": len(all_tasks),
        "todo": sum(1 for task in all_tasks if task.status == "todo"),
        "doing": sum(1 for task in all_tasks if task.status == "doing"),
        "done": sum(1 for task in all_tasks if task.status == "done"),
        "overdue": sum(1 for task in all_tasks if task.is_overdue),
        "due_soon": sum(1 for task in all_tasks if task.is_due_soon),
    }
    columns = {
        "todo": [task for task in tasks if task.status == "todo"],
        "doing": [task for task in tasks if task.status == "doing"],
        "done": [task for task in tasks if task.status == "done"],
    }
    return render_template(
        "tasks/index.html",
        tasks=tasks,
        columns=columns,
        stats=stats,
        assignees=assignees,
        category_labels=CATEGORY_LABELS,
        status_labels=STATUS_LABELS,
        priority_labels=PRIORITY_LABELS,
        q=q,
    )


@tasks_bp.route("/new", methods=["GET", "POST"])
@login_required
def create():
    wedding = current_wedding()
    form = TaskForm()
    configure_vendor_choices(form, wedding.id)
    if form.validate_on_submit():
        task = Task(wedding_id=wedding.id)
        apply_form(task, form)
        db.session.add(task)
        db.session.flush()
        audit(wedding.id, task, "create", f"נוספה המשימה {task.title}")
        db.session.commit()
        flash("המשימה נוספה.", "success")
        return redirect(url_for("tasks.index"))
    return render_template("tasks/form.html", form=form, title="הוספת משימה")


@tasks_bp.route("/<int:task_id>/edit", methods=["GET", "POST"])
@login_required
def edit(task_id: int):
    wedding = current_wedding()
    task = db.get_or_404(Task, task_id)
    if task.wedding_id != wedding.id or task.is_deleted:
        abort(404)
    form = TaskForm(obj=task)
    configure_vendor_choices(form, wedding.id)
    if request.method == "GET":
        form.related_vendor_id.data = task.related_vendor_id or 0
    if form.validate_on_submit():
        apply_form(task, form)
        audit(wedding.id, task, "update", f"עודכנה המשימה {task.title}")
        db.session.commit()
        flash("המשימה עודכנה.", "success")
        return redirect(url_for("tasks.index"))
    return render_template("tasks/form.html", form=form, title=f"עריכת {task.title}", task=task)


@tasks_bp.post("/<int:task_id>/status/<status>")
@login_required
def change_status(task_id: int, status: str):
    if status not in STATUS_LABELS:
        abort(400)
    wedding = current_wedding()
    task = db.get_or_404(Task, task_id)
    if task.wedding_id != wedding.id or task.is_deleted:
        abort(404)
    task.set_status(status)
    audit(wedding.id, task, "status", f"סטטוס המשימה {task.title} שונה ל{STATUS_LABELS[status]}")
    db.session.commit()
    flash("סטטוס המשימה עודכן.", "success")
    return redirect(request.referrer or url_for("tasks.index"))


@tasks_bp.post("/<int:task_id>/delete")
@login_required
def delete(task_id: int):
    wedding = current_wedding()
    task = db.get_or_404(Task, task_id)
    if task.wedding_id != wedding.id:
        abort(404)
    task.soft_delete()
    audit(wedding.id, task, "delete", f"המשימה {task.title} הועברה לסל המחזור")
    db.session.commit()
    flash("המשימה הועברה לסל המחזור.", "success")
    return redirect(url_for("tasks.index"))


@tasks_bp.get("/export.xlsx")
@login_required
def export_excel():
    wedding = current_wedding()
    tasks = db.session.scalars(
        db.select(Task)
        .where(Task.wedding_id == wedding.id, Task.deleted_at.is_(None))
        .order_by(Task.status, Task.due_date, Task.title)
    ).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "משימות"
    ws.sheet_view.rightToLeft = True
    headers = [
        "משימה",
        "קטגוריה",
        "סטטוס",
        "עדיפות",
        "תאריך יעד",
        "אחראי",
        "ספק קשור",
        "באיחור",
        "הערות",
    ]
    ws.append(headers)
    style_header(ws)
    for task in tasks:
        ws.append(
            [
                task.title,
                CATEGORY_LABELS.get(task.category, task.category),
                STATUS_LABELS.get(task.status, task.status),
                PRIORITY_LABELS.get(task.priority, task.priority),
                task.due_date.isoformat() if task.due_date else "",
                task.assigned_to or "",
                task.related_vendor.name if task.related_vendor else "",
                "כן" if task.is_overdue else "לא",
                task.notes or "",
            ]
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [34, 16, 14, 12, 14, 18, 22, 11, 42]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    summary = wb.create_sheet("סיכום")
    summary.sheet_view.rightToLeft = True
    summary.append(["מדד", "ערך"])
    style_header(summary)
    summary.append(["סה״כ משימות", len(tasks)])
    summary.append(["לביצוע", sum(1 for task in tasks if task.status == "todo")])
    summary.append(["בתהליך", sum(1 for task in tasks if task.status == "doing")])
    summary.append(["הושלמו", sum(1 for task in tasks if task.status == "done")])
    summary.append(["באיחור", sum(1 for task in tasks if task.is_overdue)])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="wedding-tasks.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@tasks_bp.get("/share")
@login_required
def share_text():
    wedding = current_wedding()
    tasks = db.session.scalars(
        db.select(Task)
        .where(Task.wedding_id == wedding.id, Task.deleted_at.is_(None), Task.status != "done")
        .order_by(Task.due_date.is_(None), Task.due_date, Task.priority.desc(), Task.title)
    ).all()
    lines = ["📋 משימות פתוחות", ""]
    for task in tasks:
        due = f" — עד {task.due_date.strftime('%d/%m')}" if task.due_date else ""
        owner = f" ({task.assigned_to})" if task.assigned_to else ""
        urgent = "❗ " if task.priority in {"urgent", "high"} else ""
        lines.append(f"☐ {urgent}{task.title}{owner}{due}")
    if not tasks:
        lines.append("כל המשימות הושלמו 🎉")
    text = "\n".join(lines)
    return render_template(
        "tasks/share.html", text=text, whatsapp_url=f"https://wa.me/?text={quote(text)}"
    )


def apply_form(task: Task, form: TaskForm) -> None:
    task.title = form.title.data.strip()
    task.category = form.category.data
    task.priority = form.priority.data
    task.due_date = form.due_date.data
    task.assigned_to = (form.assigned_to.data or "").strip() or None
    task.related_vendor_id = form.related_vendor_id.data or None
    task.notes = (form.notes.data or "").strip() or None
    task.set_status(form.status.data)


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="40513B")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
