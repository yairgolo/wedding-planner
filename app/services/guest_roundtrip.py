from __future__ import annotations

import csv
from collections.abc import Iterable
from dataclasses import dataclass, field
from io import StringIO
from typing import Any

from openpyxl import load_workbook

from app.extensions import db
from app.models import Family, Guest

EXPORT_SCHEMA_VERSION = "wedding-planner-guests-v1"

# These headers are intentionally stable. Exported files can always be imported back.
HEADERS = [
    "מזהה מערכת",
    "UUID",
    "שם פרטי",
    "שם משפחה",
    "טלפון",
    "אימייל",
    "צד",
    "קבוצה",
    "משפחה",
    "הוזמנו",
    "אישרו",
    "סטטוס",
    "VIP",
    "תזונה",
    "הערות תזונה",
    "שולחן",
    "הזמנה נשלחה",
    "הערות",
]

SIDE_TO_LABEL = {"groom": "חתן", "bride": "כלה", "shared": "משותף"}
LABEL_TO_SIDE = {
    "חתן": "groom",
    "כלה": "bride",
    "משותף": "shared",
    "groom": "groom",
    "bride": "bride",
    "shared": "shared",
}
STATUS_TO_LABEL = {
    "pending": "ממתין",
    "confirmed": "מגיע",
    "declined": "לא מגיע",
    "maybe": "אולי",
}
LABEL_TO_STATUS = {
    "ממתין": "pending",
    "מגיע": "confirmed",
    "לא מגיע": "declined",
    "אולי": "maybe",
    "pending": "pending",
    "confirmed": "confirmed",
    "declined": "declined",
    "maybe": "maybe",
}
DIET_TO_LABEL = {
    "regular": "רגיל",
    "vegetarian": "צמחוני",
    "vegan": "טבעוני",
    "gluten_free": "ללא גלוטן",
    "child": "מנת ילדים",
    "other": "אחר",
}
LABEL_TO_DIET = {
    "רגיל": "regular",
    "צמחוני": "vegetarian",
    "טבעוני": "vegan",
    "ללא גלוטן": "gluten_free",
    "מנת ילדים": "child",
    "אחר": "other",
    "regular": "regular",
    "vegetarian": "vegetarian",
    "vegan": "vegan",
    "gluten_free": "gluten_free",
    "child": "child",
    "other": "other",
}

ALIASES = {
    "מזהה מערכת": ("מזהה מערכת", "system_id", "id"),
    "UUID": ("uuid", "guest_uuid", "מזהה מוזמן"),
    "שם פרטי": ("שם פרטי", "first_name", "firstname"),
    "שם משפחה": ("שם משפחה", "last_name", "lastname"),
    "טלפון": ("טלפון", "phone", "mobile"),
    "אימייל": ("אימייל", "email", "mail"),
    "צד": ("צד", "side"),
    "קבוצה": ("קבוצה", "group", "group_name"),
    "משפחה": ("משפחה", "family", "family_name"),
    "הוזמנו": ("הוזמנו", "כמות", "invited_count"),
    "אישרו": ("אישרו", "confirmed_count"),
    "סטטוס": ("סטטוס", "status", "rsvp_status"),
    "VIP": ("vip", "VIP"),
    "תזונה": ("תזונה", "diet"),
    "הערות תזונה": ("הערות תזונה", "diet_notes"),
    "שולחן": ("שולחן", "table", "table_number"),
    "הזמנה נשלחה": ("הזמנה נשלחה", "invitation_sent"),
    "הערות": ("הערות", "notes"),
}


@dataclass
class ImportWarning:
    row_number: int
    message: str


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    warnings: list[ImportWarning] = field(default_factory=list)

    @property
    def warning_count(self) -> int:
        return len(self.warnings)


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalized_key(value: Any) -> str:
    return _text(value).lower().replace("_", " ").strip()


def _lookup(row: dict[str, Any], canonical: str) -> Any:
    normalized = {_normalized_key(k): v for k, v in row.items() if k is not None}
    for alias in ALIASES[canonical]:
        key = _normalized_key(alias)
        if key in normalized:
            return normalized[key]
    return None


def _integer(
    value: Any,
    default: int,
    minimum: int,
    maximum: int,
    warnings: list[str],
    label: str,
) -> int:
    if value in (None, ""):
        return default
    try:
        number = int(float(str(value).replace(",", "").strip()))
    except (TypeError, ValueError):
        warnings.append(f"{label} לא חוקי; הוגדר ל־{default}")
        return default
    if number < minimum:
        warnings.append(f"{label} קטן מדי; הוגדר ל־{minimum}")
        return minimum
    if number > maximum:
        warnings.append(f"{label} גדול מדי; הוגדר ל־{maximum}")
        return maximum
    return number


def _boolean(value: Any, default: bool = False) -> bool:
    text = _text(value).lower()
    if text in {"כן", "true", "1", "yes", "y"}:
        return True
    if text in {"לא", "false", "0", "no", "n"}:
        return False
    return default


def clean_phone(value: Any) -> str | None:
    cleaned = "".join(ch for ch in _text(value) if ch.isdigit() or ch == "+")
    return cleaned or None


def guest_to_row(guest: Guest) -> list[Any]:
    return [
        guest.id,
        guest.uuid,
        guest.first_name,
        guest.last_name or "",
        guest.phone or "",
        guest.email or "",
        SIDE_TO_LABEL.get(guest.side, guest.side),
        guest.group_name or "",
        guest.family.name if guest.family else "",
        guest.invited_count,
        guest.confirmed_count,
        STATUS_TO_LABEL.get(guest.rsvp_status, guest.rsvp_status),
        "כן" if guest.is_vip else "לא",
        DIET_TO_LABEL.get(guest.diet, guest.diet),
        guest.diet_notes or "",
        guest.table_number or "",
        "כן" if guest.invitation_sent else "לא",
        guest.notes or "",
    ]


def read_uploaded_rows(upload) -> list[dict[str, Any]]:
    filename = (upload.filename or "").lower()
    if filename.endswith(".xlsx"):
        workbook = load_workbook(upload, read_only=True, data_only=True)
        worksheet = workbook["מוזמנים"] if "מוזמנים" in workbook.sheetnames else workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        headers = next(iterator, ())
        return [dict(zip(headers, values, strict=False)) for values in iterator]

    raw = upload.stream.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("windows-1255", errors="replace")
    return list(csv.DictReader(StringIO(text)))


def _family_for_name(wedding_id: int, name: str) -> Family | None:
    if not name:
        return None
    family = db.session.scalar(
        db.select(Family).where(Family.wedding_id == wedding_id, Family.name == name)
    )
    if family:
        return family
    family = Family(wedding_id=wedding_id, name=name)
    db.session.add(family)
    db.session.flush()
    return family


def _find_existing(wedding_id: int, row: dict[str, Any], duplicate_mode: str) -> Guest | None:
    uuid_value = _text(_lookup(row, "UUID"))
    if uuid_value:
        existing = db.session.scalar(
            db.select(Guest).where(Guest.wedding_id == wedding_id, Guest.uuid == uuid_value)
        )
        if existing:
            return existing

    system_id = _text(_lookup(row, "מזהה מערכת"))
    if system_id.isdigit():
        existing = db.session.get(Guest, int(system_id))
        if existing and existing.wedding_id == wedding_id:
            return existing

    if duplicate_mode in {"update", "skip"}:
        phone = clean_phone(_lookup(row, "טלפון"))
        if phone:
            return db.session.scalar(
                db.select(Guest).where(
                    Guest.wedding_id == wedding_id,
                    Guest.phone == phone,
                    Guest.deleted_at.is_(None),
                )
            )
    return None


def import_guest_rows(
    wedding_id: int,
    rows: Iterable[dict[str, Any]],
    duplicate_mode: str = "update",
) -> ImportResult:
    result = ImportResult()

    for row_number, row in enumerate(rows, start=2):
        row_warnings: list[str] = []
        first_name = _text(_lookup(row, "שם פרטי"))
        last_name = _text(_lookup(row, "שם משפחה"))

        # Support old export files with a single full-name column.
        if not first_name:
            full_name = _text(row.get("שם") or row.get("שם מלא") or row.get("name"))
            if full_name:
                parts = full_name.split(maxsplit=1)
                first_name = parts[0]
                if not last_name and len(parts) > 1:
                    last_name = parts[1]

        if not first_name:
            first_name = "ללא שם"
            row_warnings.append("חסר שם; הרשומה יובאה בשם 'ללא שם'")

        existing = _find_existing(wedding_id, row, duplicate_mode)
        if existing and duplicate_mode == "skip":
            result.skipped += 1
            continue

        guest = existing or Guest(wedding_id=wedding_id)
        guest.deleted_at = None
        guest.first_name = first_name[:100]
        guest.last_name = last_name[:100] or None
        guest.phone = clean_phone(_lookup(row, "טלפון"))
        guest.email = _text(_lookup(row, "אימייל")).lower()[:255] or None

        raw_side = _text(_lookup(row, "צד")).lower()
        guest.side = LABEL_TO_SIDE.get(
            raw_side,
            LABEL_TO_SIDE.get(_text(_lookup(row, "צד")), "shared"),
        )
        if (
            raw_side
            and raw_side not in LABEL_TO_SIDE
            and _text(_lookup(row, "צד")) not in LABEL_TO_SIDE
        ):
            row_warnings.append("ערך צד לא מוכר; הוגדר 'משותף'")

        guest.group_name = _text(_lookup(row, "קבוצה"))[:100] or None
        family = _family_for_name(wedding_id, _text(_lookup(row, "משפחה"))[:120])
        guest.family_id = family.id if family else None
        guest.invited_count = _integer(
            _lookup(row, "הוזמנו"), 1, 1, 100, row_warnings, "כמות מוזמנים"
        )
        guest.confirmed_count = _integer(
            _lookup(row, "אישרו"), 0, 0, guest.invited_count, row_warnings, "כמות מאשרים"
        )

        raw_status = _text(_lookup(row, "סטטוס"))
        guest.rsvp_status = LABEL_TO_STATUS.get(raw_status, "pending")
        if raw_status and raw_status not in LABEL_TO_STATUS:
            row_warnings.append("סטטוס לא מוכר; הוגדר 'ממתין'")

        guest.is_vip = _boolean(_lookup(row, "VIP"), False)
        raw_diet = _text(_lookup(row, "תזונה"))
        guest.diet = LABEL_TO_DIET.get(raw_diet, "regular")
        if raw_diet and raw_diet not in LABEL_TO_DIET:
            guest.diet = "other"
            row_warnings.append("סוג תזונה לא מוכר; הוגדר 'אחר'")
        guest.diet_notes = _text(_lookup(row, "הערות תזונה"))[:255] or None
        guest.table_number = _text(_lookup(row, "שולחן"))[:30] or None
        guest.invitation_sent = _boolean(_lookup(row, "הזמנה נשלחה"), guest.invitation_sent)
        guest.notes = _text(_lookup(row, "הערות")) or None

        db.session.add(guest)
        if existing:
            result.updated += 1
        else:
            result.created += 1

        for warning in row_warnings:
            result.warnings.append(ImportWarning(row_number=row_number, message=warning))

    db.session.commit()
    return result
