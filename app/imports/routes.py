import csv
from io import StringIO

from flask import Blueprint, flash, redirect, render_template, url_for
from flask_login import login_required
from openpyxl import load_workbook

from app.extensions import db
from app.models import Guest, Wedding
from app.services.audit import log_action

from .forms import GuestImportForm

imports_bp = Blueprint("imports", __name__, url_prefix="/imports")


def clean_phone(value):
    return "".join(ch for ch in str(value or "") if ch.isdigit() or ch == "+")


def normalize_row(row):
    keys = {str(k).strip().lower(): v for k, v in row.items() if k is not None}
    name = keys.get("שם") or keys.get("name") or keys.get("שם מלא") or ""
    first = keys.get("שם פרטי") or keys.get("first_name") or name
    last = keys.get("שם משפחה") or keys.get("last_name") or ""
    return {
        "first_name": str(first or "").strip(),
        "last_name": str(last or "").strip(),
        "phone": clean_phone(keys.get("טלפון") or keys.get("phone")),
        "email": str(keys.get("אימייל") or keys.get("email") or "").strip(),
        "side": str(keys.get("צד") or keys.get("side") or "shared").strip().lower(),
        "group_name": str(keys.get("קבוצה") or keys.get("group") or "").strip(),
        "invited_count": int(keys.get("כמות") or keys.get("invited_count") or 1),
    }


@imports_bp.route("", methods=["GET", "POST"])
@login_required
def index():
    form = GuestImportForm()
    if form.validate_on_submit():
        wedding = db.session.scalar(db.select(Wedding).order_by(Wedding.id).limit(1))
        upload = form.file.data
        rows = []
        if upload.filename.lower().endswith(".xlsx"):
            wb = load_workbook(upload, read_only=True, data_only=True)
            ws = wb.active
            values = list(ws.iter_rows(values_only=True))
            headers = values[0] if values else []
            rows = [dict(zip(headers, row, strict=False)) for row in values[1:]]
        else:
            text = upload.stream.read().decode("utf-8-sig")
            rows = list(csv.DictReader(StringIO(text)))
        imported = skipped = 0
        for raw in rows:
            data = normalize_row(raw)
            if not data["first_name"]:
                continue
            if form.duplicate_mode.data == "skip" and data["phone"]:
                exists = db.session.scalar(
                    db.select(Guest).where(
                        Guest.wedding_id == wedding.id,
                        Guest.phone == data["phone"],
                        Guest.deleted_at.is_(None),
                    )
                )
                if exists:
                    skipped += 1
                    continue
            db.session.add(Guest(wedding_id=wedding.id, **data))
            imported += 1
        db.session.commit()
        log_action("guest_import", None, "imported", f"יובאו {imported} מוזמנים; דולגו {skipped}")
        flash(f"יובאו {imported} מוזמנים. דולגו {skipped} כפילויות.", "success")
        return redirect(url_for("guests.index"))
    return render_template("imports/index.html", form=form)
