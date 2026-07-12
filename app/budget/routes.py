from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from urllib.parse import quote

from flask import Blueprint, abort, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import or_

from app.extensions import db
from app.models import AuditLog, BudgetItem, ShoppingItem, Wedding

from .forms import BudgetItemForm, BudgetTargetForm

budget_bp = Blueprint("budget", __name__, url_prefix="/budget")

CATEGORY_LABELS = {
    "venue": "אולם ואוכל",
    "photography": "צילום",
    "music": "מוזיקה ו-DJ",
    "clothing": "לבוש",
    "beauty": "איפור ושיער",
    "design": "עיצוב והזמנות",
    "attractions": "אטרקציות",
    "transport": "הסעות ורכב",
    "home": "בית",
    "other": "אחר",
}
STATUS_LABELS = {
    "planned": "מתוכנן",
    "agreed": "נסגר",
    "partial": "שולם חלקית",
    "paid": "שולם",
    "cancelled": "בוטל",
}


def current_wedding() -> Wedding:
    wedding = db.session.scalar(db.select(Wedding).order_by(Wedding.id).limit(1))
    if not wedding:
        abort(404)
    return wedding


def audit(wedding_id: int, item: BudgetItem, action: str, description: str) -> None:
    db.session.add(
        AuditLog(
            wedding_id=wedding_id,
            user_id=current_user.id,
            entity_type="budget_item",
            entity_id=str(item.id or "new"),
            action=action,
            description=description,
        )
    )


@budget_bp.route("", methods=["GET", "POST"])
@login_required
def index():
    wedding = current_wedding()
    target_form = BudgetTargetForm(obj=wedding)
    if target_form.validate_on_submit() and target_form.submit.data:
        wedding.budget_target = target_form.budget_target.data or 0
        db.session.add(
            AuditLog(
                wedding_id=wedding.id,
                user_id=current_user.id,
                entity_type="wedding",
                entity_id=str(wedding.id),
                action="budget_target",
                description=f"תקציב היעד עודכן ל-₪{wedding.budget_target}",
            )
        )
        db.session.commit()
        flash("תקציב היעד עודכן.", "success")
        return redirect(url_for("budget.index"))

    q = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    status = request.args.get("status", "").strip()
    stmt = db.select(BudgetItem).where(
        BudgetItem.wedding_id == wedding.id, BudgetItem.deleted_at.is_(None)
    )
    if q:
        pattern = f"%{q}%"
        stmt = stmt.where(
            or_(
                BudgetItem.name.ilike(pattern),
                BudgetItem.supplier_name.ilike(pattern),
                BudgetItem.notes.ilike(pattern),
            )
        )
    if category:
        stmt = stmt.where(BudgetItem.category == category)
    if status:
        stmt = stmt.where(BudgetItem.status == status)
    items = db.session.scalars(
        stmt.order_by(BudgetItem.due_date, BudgetItem.created_at.desc())
    ).all()
    all_items = db.session.scalars(
        db.select(BudgetItem).where(
            BudgetItem.wedding_id == wedding.id, BudgetItem.deleted_at.is_(None)
        )
    ).all()
    shopping_actual = db.session.scalars(
        db.select(ShoppingItem).where(
            ShoppingItem.wedding_id == wedding.id,
            ShoppingItem.deleted_at.is_(None),
            ShoppingItem.status == "purchased",
        )
    ).all()
    committed = sum(
        (item.committed_amount for item in all_items if item.status != "cancelled"),
        Decimal("0"),
    )
    paid = sum(
        (item.paid_amount or Decimal("0") for item in all_items if item.status != "cancelled"),
        Decimal("0"),
    )
    shopping_total = sum((item.total_actual for item in shopping_actual), Decimal("0"))
    target = wedding.budget_target or Decimal("0")
    total_with_shopping = committed + shopping_total
    stats = {
        "target": target,
        "committed": committed,
        "paid": paid,
        "balance": max(Decimal("0"), committed - paid),
        "shopping": shopping_total,
        "total": total_with_shopping,
        "remaining": target - total_with_shopping,
        "percent": min(100, int((total_with_shopping / target) * 100)) if target > 0 else 0,
        "overdue": sum(1 for item in all_items if item.is_overdue),
    }
    return render_template(
        "budget/index.html",
        wedding=wedding,
        items=items,
        stats=stats,
        target_form=target_form,
        category_labels=CATEGORY_LABELS,
        status_labels=STATUS_LABELS,
        q=q,
    )


@budget_bp.route("/new", methods=["GET", "POST"])
@login_required
def create():
    wedding = current_wedding()
    form = BudgetItemForm()
    if form.validate_on_submit():
        item = BudgetItem(wedding_id=wedding.id)
        apply_form(item, form)
        normalize_status(item)
        db.session.add(item)
        db.session.flush()
        audit(wedding.id, item, "create", f"נוספה הוצאה {item.name}")
        db.session.commit()
        flash("ההוצאה נוספה לתקציב.", "success")
        return redirect(url_for("budget.index"))
    return render_template("budget/form.html", form=form, title="הוספת הוצאה")


@budget_bp.route("/<int:item_id>/edit", methods=["GET", "POST"])
@login_required
def edit(item_id: int):
    wedding = current_wedding()
    item = db.get_or_404(BudgetItem, item_id)
    if item.wedding_id != wedding.id or item.is_deleted:
        abort(404)
    form = BudgetItemForm(obj=item)
    if form.validate_on_submit():
        apply_form(item, form)
        normalize_status(item)
        audit(wedding.id, item, "update", f"עודכנה ההוצאה {item.name}")
        db.session.commit()
        flash("ההוצאה עודכנה.", "success")
        return redirect(url_for("budget.index"))
    return render_template("budget/form.html", form=form, title=f"עריכת {item.name}", item=item)


@budget_bp.post("/<int:item_id>/paid")
@login_required
def mark_paid(item_id: int):
    wedding = current_wedding()
    item = db.get_or_404(BudgetItem, item_id)
    if item.wedding_id != wedding.id or item.is_deleted:
        abort(404)
    item.paid_amount = item.committed_amount
    item.status = "paid"
    audit(wedding.id, item, "paid", f"ההוצאה {item.name} סומנה כשולמה")
    db.session.commit()
    flash("התשלום סומן כמלא.", "success")
    return redirect(request.referrer or url_for("budget.index"))


@budget_bp.post("/<int:item_id>/delete")
@login_required
def delete(item_id: int):
    wedding = current_wedding()
    item = db.get_or_404(BudgetItem, item_id)
    if item.wedding_id != wedding.id:
        abort(404)
    item.soft_delete()
    audit(wedding.id, item, "delete", f"ההוצאה {item.name} הועברה לסל המחזור")
    db.session.commit()
    flash("ההוצאה הועברה לסל המחזור.", "success")
    return redirect(url_for("budget.index"))


@budget_bp.get("/export.xlsx")
@login_required
def export_excel():
    wedding = current_wedding()
    items = db.session.scalars(
        db.select(BudgetItem)
        .where(BudgetItem.wedding_id == wedding.id, BudgetItem.deleted_at.is_(None))
        .order_by(BudgetItem.category, BudgetItem.name)
    ).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "תקציב"
    ws.sheet_view.rightToLeft = True
    ws.append(
        [
            "הוצאה",
            "קטגוריה",
            "ספק",
            "מתוכנן",
            "סכום בפועל",
            "מחויבות",
            "שולם",
            "נותר",
            "סטטוס",
            "תשלום הבא",
            "הערות",
        ]
    )
    style_header(ws)
    for item in items:
        ws.append(
            [
                item.name,
                CATEGORY_LABELS.get(item.category, item.category),
                item.supplier_name or "",
                float(item.planned_amount or 0),
                float(item.actual_amount or 0),
                float(item.committed_amount),
                float(item.paid_amount or 0),
                float(item.balance),
                STATUS_LABELS.get(item.status, item.status),
                item.due_date.isoformat() if item.due_date else "",
                item.notes or "",
            ]
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, width in enumerate([28, 18, 22, 14, 14, 14, 14, 14, 15, 16, 34], 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    summary = wb.create_sheet("סיכום")
    summary.sheet_view.rightToLeft = True
    summary.append(["מדד", "סכום"])
    style_header(summary)
    committed = sum(float(item.committed_amount) for item in items if item.status != "cancelled")
    paid = sum(float(item.paid_amount or 0) for item in items if item.status != "cancelled")
    summary.append(["תקציב יעד", float(wedding.budget_target or 0)])
    summary.append(["התחייבויות", committed])
    summary.append(["שולם", paid])
    summary.append(["נותר לתשלום", committed - paid])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="wedding-budget.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@budget_bp.get("/share")
@login_required
def share_text():
    wedding = current_wedding()
    items = db.session.scalars(
        db.select(BudgetItem)
        .where(
            BudgetItem.wedding_id == wedding.id,
            BudgetItem.deleted_at.is_(None),
            BudgetItem.status != "cancelled",
        )
        .order_by(BudgetItem.due_date, BudgetItem.name)
    ).all()
    committed = sum((item.committed_amount for item in items), Decimal("0"))
    paid = sum((item.paid_amount or Decimal("0") for item in items), Decimal("0"))
    lines = [
        "💰 סיכום תקציב חתונה",
        "",
        f"תקציב יעד: ₪{float(wedding.budget_target or 0):,.0f}",
        f"התחייבויות: ₪{float(committed):,.0f}",
        f"שולם: ₪{float(paid):,.0f}",
        f"נותר לתשלום: ₪{float(committed - paid):,.0f}",
        "",
        "*תשלומים פתוחים:*",
    ]
    for item in items:
        if item.balance > 0:
            due = f" · {item.due_date.strftime('%d/%m/%Y')}" if item.due_date else ""
            lines.append(f"• {item.name}: ₪{float(item.balance):,.0f}{due}")
    text = "\n".join(lines)
    return render_template(
        "budget/share.html", text=text, whatsapp_url=f"https://wa.me/?text={quote(text)}"
    )


def apply_form(item: BudgetItem, form: BudgetItemForm) -> None:
    item.name = form.name.data.strip()
    item.category = form.category.data
    item.supplier_name = (form.supplier_name.data or "").strip() or None
    item.planned_amount = form.planned_amount.data or 0
    item.actual_amount = form.actual_amount.data or 0
    item.paid_amount = form.paid_amount.data or 0
    item.status = form.status.data
    item.due_date = form.due_date.data
    item.notes = (form.notes.data or "").strip() or None


def normalize_status(item: BudgetItem) -> None:
    if item.status == "cancelled":
        return
    if item.committed_amount > 0 and item.paid_amount >= item.committed_amount:
        item.status = "paid"
    elif item.paid_amount > 0:
        item.status = "partial"


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="40513B")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
