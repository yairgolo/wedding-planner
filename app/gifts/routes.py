from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

from flask import Blueprint, abort, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import or_

from app.extensions import db
from app.models import AuditLog, Gift, Wedding

from .forms import GiftForm

gifts_bp = Blueprint("gifts", __name__, url_prefix="/gifts")
TYPE_LABELS = {
    "cash": "מזומן",
    "check": "צ׳ק",
    "transfer": "העברה",
    "item": "מתנה פיזית",
    "other": "אחר",
}


def current_wedding():
    wedding = db.session.scalar(db.select(Wedding).order_by(Wedding.id).limit(1))
    if not wedding:
        abort(404)
    return wedding


def audit(wedding_id, gift, action, text):
    db.session.add(
        AuditLog(
            wedding_id=wedding_id,
            user_id=current_user.id,
            entity_type="gift",
            entity_id=str(gift.id or "new"),
            action=action,
            description=text,
        )
    )


@gifts_bp.get("")
@login_required
def index():
    wedding = current_wedding()
    q = request.args.get("q", "").strip()
    pending = request.args.get("pending", "")
    stmt = db.select(Gift).where(Gift.wedding_id == wedding.id, Gift.deleted_at.is_(None))
    if q:
        pattern = f"%{q}%"
        stmt = stmt.where(
            or_(
                Gift.guest_name.ilike(pattern),
                Gift.description.ilike(pattern),
                Gift.notes.ilike(pattern),
            )
        )
    if pending == "1":
        stmt = stmt.where(Gift.thank_you_sent.is_(False))
    gifts = db.session.scalars(stmt.order_by(Gift.received_date.desc(), Gift.id.desc())).all()
    all_gifts = db.session.scalars(
        db.select(Gift).where(Gift.wedding_id == wedding.id, Gift.deleted_at.is_(None))
    ).all()
    stats = {
        "count": len(all_gifts),
        "total": sum(float(x.value) for x in all_gifts),
        "pending": sum(1 for x in all_gifts if not x.thank_you_sent),
        "thanked": sum(1 for x in all_gifts if x.thank_you_sent),
    }
    return render_template(
        "gifts/index.html", gifts=gifts, stats=stats, type_labels=TYPE_LABELS, q=q
    )


@gifts_bp.route("/new", methods=["GET", "POST"])
@login_required
def create():
    wedding = current_wedding()
    form = GiftForm()
    if form.validate_on_submit():
        gift = Gift(wedding_id=wedding.id)
        apply_form(gift, form)
        db.session.add(gift)
        db.session.flush()
        audit(wedding.id, gift, "create", f"נוספה מתנה מאת {gift.guest_name}")
        db.session.commit()
        flash("המתנה נוספה.", "success")
        return redirect(url_for("gifts.index"))
    return render_template("gifts/form.html", form=form, title="הוספת מתנה")


@gifts_bp.route("/<int:gift_id>/edit", methods=["GET", "POST"])
@login_required
def edit(gift_id):
    wedding = current_wedding()
    gift = db.get_or_404(Gift, gift_id)
    if gift.wedding_id != wedding.id or gift.is_deleted:
        abort(404)
    form = GiftForm(obj=gift)
    if form.validate_on_submit():
        apply_form(gift, form)
        audit(wedding.id, gift, "update", f"עודכנה המתנה מאת {gift.guest_name}")
        db.session.commit()
        flash("המתנה עודכנה.", "success")
        return redirect(url_for("gifts.index"))
    return render_template("gifts/form.html", form=form, title=f"עריכת מתנה — {gift.guest_name}")


@gifts_bp.post("/<int:gift_id>/thank")
@login_required
def thank(gift_id):
    wedding = current_wedding()
    gift = db.get_or_404(Gift, gift_id)
    if gift.wedding_id != wedding.id or gift.is_deleted:
        abort(404)
    gift.thank_you_sent = not gift.thank_you_sent
    gift.thank_you_sent_at = datetime.now(timezone.utc) if gift.thank_you_sent else None
    audit(wedding.id, gift, "thank", f"עודכן סטטוס תודה עבור {gift.guest_name}")
    db.session.commit()
    flash("סטטוס התודה עודכן.", "success")
    return redirect(request.referrer or url_for("gifts.index"))


@gifts_bp.post("/<int:gift_id>/delete")
@login_required
def delete(gift_id):
    wedding = current_wedding()
    gift = db.get_or_404(Gift, gift_id)
    if gift.wedding_id != wedding.id:
        abort(404)
    gift.soft_delete()
    audit(wedding.id, gift, "delete", f"המתנה מאת {gift.guest_name} הועברה לסל המחזור")
    db.session.commit()
    flash("המתנה הועברה לסל המחזור.", "success")
    return redirect(url_for("gifts.index"))


@gifts_bp.get("/share")
@login_required
def share():
    wedding = current_wedding()
    gifts = db.session.scalars(
        db.select(Gift)
        .where(Gift.wedding_id == wedding.id, Gift.deleted_at.is_(None))
        .order_by(Gift.guest_name)
    ).all()
    lines = ["🎁 סיכום מתנות", ""]
    for gift in gifts:
        gift_type = TYPE_LABELS.get(gift.gift_type, gift.gift_type)
        thank_status = "✅ תודה נשלחה" if gift.thank_you_sent else "⏳ תודה ממתינה"
        lines.append(f"{gift.guest_name} — ₪{float(gift.value):,.0f} ({gift_type}) {thank_status}")
    return render_template("gifts/share.html", share_text="\n".join(lines))


@gifts_bp.get("/export.xlsx")
@login_required
def export_excel():
    wedding = current_wedding()
    gifts = db.session.scalars(
        db.select(Gift)
        .where(Gift.wedding_id == wedding.id, Gift.deleted_at.is_(None))
        .order_by(Gift.guest_name)
    ).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "מתנות"
    ws.sheet_view.rightToLeft = True
    headers = ["שם", "סוג", "סכום / שווי", "תיאור", "תאריך", "תודה נשלחה", "הערות"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="40513B")
        c.alignment = Alignment(horizontal="center")
    for g in gifts:
        ws.append(
            [
                g.guest_name,
                TYPE_LABELS.get(g.gift_type, g.gift_type),
                float(g.value),
                g.description or "",
                g.received_date.isoformat() if g.received_date else "",
                "כן" if g.thank_you_sent else "לא",
                g.notes or "",
            ]
        )
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for i, w in enumerate([24, 16, 16, 34, 14, 16, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        as_attachment=True,
        download_name="wedding-gifts.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def apply_form(gift, form):
    gift.guest_name = form.guest_name.data.strip()
    gift.gift_type = form.gift_type.data
    gift.amount = form.amount.data or 0
    gift.description = (form.description.data or "").strip() or None
    gift.received_date = form.received_date.data
    gift.thank_you_sent = bool(form.thank_you_sent.data)
    gift.thank_you_sent_at = (
        datetime.now(timezone.utc)
        if gift.thank_you_sent and not gift.thank_you_sent_at
        else (gift.thank_you_sent_at if gift.thank_you_sent else None)
    )
    gift.notes = (form.notes.data or "").strip() or None
